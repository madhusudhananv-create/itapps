import threading
import time
import io
import datetime as dt
import os
from flask import Flask, jsonify, send_from_directory, request, send_file

import config
import compliance

app = Flask(__name__, static_folder="static", static_url_path="")
app.secret_key = config.FLASK_SECRET_KEY
APP_ROOT = os.path.dirname(os.path.abspath(__file__))


@app.after_request
def fix_windows_mime_types(response):
    """Some Windows machines have a corrupted file-type registry that makes
    Python's mimetypes module (which Flask relies on for static files) guess
    the wrong Content-Type for .css/.js — often text/plain instead of
    text/css. Browsers silently refuse to APPLY a stylesheet with the wrong
    Content-Type (even though the raw text loads fine if you visit the URL
    directly), which looks exactly like "the CSS isn't working" without any
    other visible error. Force the correct type here so this never depends
    on whatever that machine's registry happens to say."""
    path = request.path
    if path.endswith(".css"):
        response.headers["Content-Type"] = "text/css; charset=utf-8"
    elif path.endswith(".js"):
        response.headers["Content-Type"] = "application/javascript; charset=utf-8"
    return response

# ---------------------------------------------------------------------------
# Authentication is now handled entirely by the IT Apps Portal (login.html /
# homepage / microapps.html at the ITAPPS root) — this app no longer runs its
# own login flow. The frontend (static/index.html) checks for the portal's
# token in localStorage on load and redirects to the portal's /login if it's
# missing; there's no server-side gate here yet (matching the portal's own
# current approach of trusting the client-side check). If/when an allow-list
# restriction is added later, that's the place to add real enforcement.
# ---------------------------------------------------------------------------


# ---------------------------------------------------------------------------
# All 5 reports are re-read and recomputed by a single background thread on
# its own schedule, instead of inline during an HTTP request. Querying 5
# Azure Storage tables and recomputing compliance genuinely takes a few
# seconds — doing that while a browser tab is sitting there waiting is a bad
# experience, and it's unnecessary anyway since the tables only change every
# ~5 minutes. Every request below just reads whatever is already sitting in
# `_cache` — instant, no blocking — while the background thread keeps that
# cache up to date.
# ---------------------------------------------------------------------------
_cache = {
    "mtimes": None,
    "data": None,
    "computed_at": None,
    "error": None,
    "refreshing": False,
}
_cache_lock = threading.Lock()
_force_refresh_event = threading.Event()

# How often the background thread checks the files for changes. Independent
# from POLL_INTERVAL_MS (how often the browser asks the server for data).
BACKGROUND_CHECK_SECONDS = 15


def _refresh_once(force=False):
    """Recompute if a source table's watermark changed since last read (or if
    force=True). Runs on the background thread — never on a request thread."""
    current_mtimes = compliance.source_mtimes()
    with _cache_lock:
        unchanged = _cache["data"] is not None and current_mtimes == _cache["mtimes"]
    if not force and unchanged:
        return

    with _cache_lock:
        _cache["refreshing"] = True
    try:
        data = compliance.compute_dashboard()
        with _cache_lock:
            _cache.update(mtimes=current_mtimes, data=data, computed_at=time.time(),
                          error=None, refreshing=False)
    except Exception as exc:
        with _cache_lock:
            _cache.update(mtimes=current_mtimes, data=None, computed_at=time.time(),
                          error=str(exc), refreshing=False)
        print(f"[app] background refresh failed: {exc}", flush=True)


def _background_loop():
    # Compute once immediately on startup so the first page load has data.
    _refresh_once(force=True)
    while True:
        # Wait for either the check interval to elapse, or someone to hit
        # "Refresh now" (which sets _force_refresh_event early).
        triggered = _force_refresh_event.wait(timeout=BACKGROUND_CHECK_SECONDS)
        if triggered:
            _force_refresh_event.clear()
            _refresh_once(force=True)
        else:
            _refresh_once(force=False)


def start_background_thread():
    t = threading.Thread(target=_background_loop, daemon=True)
    t.start()


def _snapshot():
    with _cache_lock:
        return dict(_cache)


@app.route("/")
def index():
    return send_from_directory(app.static_folder, "index.html")


@app.route("/workmodepolicy")
@app.route("/workmodepolicy/")
@app.route("/workmodepolicy.html")
def workmodepolicy_entry():
    return send_from_directory(app.static_folder, "index.html")


@app.route("/hr")
@app.route("/hr/")
@app.route("/hr/index.html")
@app.route("/hr/workmodepolicy")
@app.route("/hr/workmodepolicy/")
@app.route("/hr/workmodepolicy.html")
def workmodepolicy_entry_prefixed():
    return send_from_directory(app.static_folder, "index.html")


@app.route("/js/<path:filename>")
def attendance_js(filename):
    return send_from_directory(os.path.join(APP_ROOT, "js"), filename)


@app.route("/hr/js/<path:filename>")
def attendance_js_prefixed(filename):
    return send_from_directory(os.path.join(APP_ROOT, "js"), filename)


@app.route("/hr/<path:filename>")
def attendance_static_prefixed(filename):
    return send_from_directory(app.static_folder, filename)


@app.route("/api/config")
def api_config():
    return jsonify({"pollIntervalMs": config.POLL_INTERVAL_MS})


@app.route("/hr/api/config")
def api_config_prefixed():
    return api_config()


def _range_for_week_offset(week_offset: int):
    """week_offset=1 -> last completed Mon-Sun calendar week,
    week_offset=2 -> the week before that, etc.
    Always anchored to today's calendar date, not to the live biometric window."""
    today = dt.date.today()
    # Monday of the current calendar week
    start_of_this_week = today - dt.timedelta(days=today.weekday())
    start = start_of_this_week - dt.timedelta(days=7 * week_offset)
    end = start + dt.timedelta(days=6)
    return start.isoformat(), end.isoformat()


# ---------------------------------------------------------------------------
# In-memory cache for historical weeks queried directly from Azure Tables.
# Keyed by week_end date string (ISO format). TTL prevents stale data.
# ---------------------------------------------------------------------------
_week_cache: dict = {}
_week_cache_lock = threading.Lock()
_week_computing: set = set()
_week_computing_lock = threading.Lock()
WEEK_CACHE_TTL_SECONDS = int(os.environ.get("WEEK_CACHE_TTL_SECONDS", 3600))


def _get_or_schedule_week(start_date: str, end_date: str):
    """Return (data, note, still_computing).

    First call triggers a background Azure query and returns still_computing=True.
    Subsequent calls within the TTL return from the in-memory cache immediately.
    A failed computation stores data=None so callers can surface an error note."""
    cache_key = f"{start_date}_{end_date}"
    now = time.time()
    with _week_cache_lock:
        entry = _week_cache.get(cache_key)
        if entry and (now - entry["computed_at"]) < WEEK_CACHE_TTL_SECONDS:
            if entry["data"] is None:
                return None, f"Could not load data for {start_date} to {end_date}.", False
            return entry["data"], None, False

    with _week_computing_lock:
        if cache_key in _week_computing:
            return None, None, True
        _week_computing.add(cache_key)

    def _compute_week():
        try:
            data = compliance.compute_dashboard(window_override=(start_date, end_date))
            with _week_cache_lock:
                _week_cache[cache_key] = {"data": data, "computed_at": time.time()}
            print(f"[app] cached week {start_date}..{end_date}: {len(data.get('employees', []))} employees", flush=True)
        except Exception as exc:
            print(f"[app] week compute failed {start_date}..{end_date}: {exc}", flush=True)
            with _week_cache_lock:
                _week_cache[cache_key] = {"data": None, "computed_at": time.time()}
        finally:
            with _week_computing_lock:
                _week_computing.discard(cache_key)

    threading.Thread(target=_compute_week, daemon=True).start()
    return None, None, True


def _get_data_for_range(start_date: str, end_date: str):
    """Query Azure Tables for an arbitrary date range via in-memory caching.
    Returns (data_dict, note). data is None and note is a loading/error
    message when the computation is in progress or has failed."""
    data, note, still_computing = _get_or_schedule_week(start_date, end_date)
    if still_computing:
        return None, f"Loading {start_date} to {end_date} from Azure — please wait a moment."
    return data, note


def _get_data_for_week(week_offset: int):
    """Returns (data, computed_at, note, still_starting).

    still_starting=True means data is being computed (either the initial
    startup computation or a historical week query) — the frontend should
    show a spinner and auto-retry."""
    if week_offset <= 0:
        snap = _snapshot()
        return snap["data"], snap["computed_at"], None, False

    start, end = _range_for_week_offset(week_offset)
    data, note, still_computing = _get_or_schedule_week(start, end)
    if still_computing:
        return None, None, f"Loading {start} to {end} from Azure — please wait a moment.", True
    if data is None:
        return None, None, note or f"Could not load data for {start} to {end}.", False
    return data, None, note or f"Showing {start} to {end}.", False


@app.route("/api/dashboard")
def api_dashboard():
    start_date = request.args.get("startDate")
    end_date = request.args.get("endDate")
    checked_at = time.time()

    if start_date and end_date:
        data, coverage_note = _get_data_for_range(start_date, end_date)
        if data is None:
            msg = coverage_note or f"Loading {start_date} to {end_date} from Azure — please wait a moment."
            return jsonify({"ok": False, "error": msg, "computedAt": None,
                             "checkedAt": checked_at, "refreshing": True}), 200
        note = coverage_note or f"Showing {start_date} to {end_date}."
        return jsonify({"ok": True, "computedAt": None, "checkedAt": checked_at,
                         "refreshing": False, "historyNote": note, **data}), 200

    week = request.args.get("week", 0, type=int)
    if week > 0:
        data, _, note, still_starting = _get_data_for_week(week)
        if data is None:
            return jsonify({"ok": False, "error": note, "computedAt": None,
                             "checkedAt": checked_at, "refreshing": still_starting}), 200
        return jsonify({"ok": True, "computedAt": None, "checkedAt": checked_at,
                         "refreshing": False, "historyNote": note, **data}), 200

    snap = _snapshot()
    if snap["error"] and snap["data"] is None:
        return jsonify({"ok": False, "error": snap["error"], "computedAt": snap["computed_at"],
                         "checkedAt": checked_at, "refreshing": snap["refreshing"]}), 200
    if snap["data"] is None:
        # First request may land before the background thread's initial run finishes.
        return jsonify({"ok": False, "error": "Still loading the reports for the first time…",
                         "computedAt": None, "checkedAt": checked_at, "refreshing": True}), 200
    return jsonify({"ok": True, "computedAt": snap["computed_at"], "checkedAt": checked_at,
                     "refreshing": snap["refreshing"], **snap["data"]}), 200


@app.route("/hr/api/dashboard")
def api_dashboard_prefixed():
    return api_dashboard()


@app.route("/api/refresh", methods=["POST"])
def api_refresh():
    """'Refresh now' nudges the background thread to recompute immediately,
    then returns whatever is currently cached right away — it does not block
    waiting for the new computation to finish. The frontend keeps polling
    /api/dashboard afterward and will pick up the new data within a second
    or two once the background thread completes."""
    _force_refresh_event.set()
    snap = _snapshot()
    checked_at = time.time()
    if snap["data"] is None:
        return jsonify({"ok": False, "error": snap["error"] or "Still loading…",
                         "computedAt": snap["computed_at"], "checkedAt": checked_at,
                         "refreshing": True}), 200
    return jsonify({"ok": True, "computedAt": snap["computed_at"], "checkedAt": checked_at,
                     "refreshing": True, **snap["data"]}), 200


@app.route("/hr/api/refresh", methods=["POST"])
def api_refresh_prefixed():
    return api_refresh()


start_background_thread()

# ---------------------------------------------------------------------------
# "View employees" / "Export to Excel" — used by every clickable number on
# the dashboard. Both routes share the same filtering logic so the exported
# file always matches exactly what the linked page shows.
# ---------------------------------------------------------------------------
FILTER_PARAM_TO_FIELD = {
    "practice": "practice",
    "subPractice": "subPractice",
    "level": "level",
    "project": "project",
    "businessUnit": "businessUnit",
    "location": "location",
    "reportingManager": "reportingManager",
    "hrbp": "hrbp",
}

CATEGORY_LABELS = {
    "tracked": "Eligible Employees",
    "compliant": "Compliant — 3+ Office Days",
    "covered": "Non-Compliant Employees",
    "flagged": "Non-Compliant Employees",
    "bucket2": "2 Days in Office",
    "bucket1": "1 Day in Office",
    "bucket0": "0 Days in Office",
    "clientLocation": "Employees Working from Client Location",
    "exception": "WFH Exception",
}


def _select_category(data, category):
    employees = data.get("employees", [])
    client_loc = data.get("clientLocationEmployees", [])
    exception_emps = data.get("exceptionEmployees", [])
    if category == "clientLocation":
        return client_loc, "simple"
    if category == "exception":
        return exception_emps, "exception"
    if category == "compliant":
        return [e for e in employees if e.get("bucket") == 3], "full"
    if category == "covered":
        return [e for e in employees if e.get("subStatus") == "covered"], "full"
    if category == "flagged":
        return [e for e in employees if e.get("subStatus") == "flagged"], "full"
    if category in ("bucket2", "bucket1", "bucket0"):
        n = int(category[-1])
        return [e for e in employees if e.get("bucket") == n], "full"
    return employees, "full"  # "tracked" or unrecognized -> everyone


def _apply_filters(employees, args):
    join_date = (args.get("joinDate") or "").strip()
    def matches(e):
        for param, field in FILTER_PARAM_TO_FIELD.items():
            val = (args.get(param) or "").strip().lower()
            if val and val not in (e.get(field) or "not available").lower():
                return False
        if join_date and (e.get("dateOfJoining") or "") < join_date:
            return False
        return True
    return [e for e in employees if matches(e)]


def _resolve_view_data():
    """Shared by /view and /export.xlsx: reads ?week= or ?startDate=/?endDate=,
    ?category=, and the filter params, and returns (rows, kind, label, error)."""
    start_date = request.args.get("startDate")
    end_date = request.args.get("endDate")
    category = request.args.get("category", "tracked")

    if start_date and end_date:
        data, note = _get_data_for_range(start_date, end_date)
        if data is None:
            return None, None, None, note or "Loading data from Azure — please wait a moment."
    else:
        week = request.args.get("week", 0, type=int)
        if week > 0:
            data, _, note, _ = _get_data_for_week(week)
            if data is None:
                return None, None, None, note
        else:
            snap = _snapshot()
            data = snap["data"]
            if data is None:
                return None, None, None, "Dashboard data isn't ready yet — try again in a moment."

    rows, kind = _select_category(data, category)
    rows = _apply_filters(rows, request.args)
    label = CATEGORY_LABELS.get(category, "Employees")
    return rows, kind, label, None


def _total_hours_from_days(days):
    total_mins = 0
    for d in days or []:
        if d.get("state") == "office":
            raw = str(d.get("hours") or "").strip()
            if ":" not in raw:
                continue
            try:
                hh, mm = raw.split(":", 1)
                total_mins += int(hh) * 60 + int(mm)
            except Exception:
                continue
    if total_mins <= 0:
        return "—"
    hh = total_mins // 60
    mm = total_mins % 60
    return f"{hh}h {mm:02d}m"


def _day_cell_html(days):
    import datetime as _dt
    letters_by_weekday = ['M','T','W','T','F','S','S']  # date.weekday(): Mon=0..Sun=6
    cells = ""
    for d in days:
        try:
            letter = letters_by_weekday[_dt.date.fromisoformat(d["date"]).weekday()]
        except Exception:
            letter = "?"
        state = d.get("state", "")
        state_label = "WFH" if state == "gap" else state
        tip_extra = ""
        if state == "office" and d.get("firstIn"):
            tip_extra = f' — in {d.get("firstIn", "")}'
            if d.get("lastOut"):
                tip_extra += f', out {d.get("lastOut", "")}'
            if d.get("hours"):
                tip_extra += f' ({d.get("hours", "")})'
        cells += f'<div class="day-cell {state}" title="{d.get("date", "")}: {state_label}{tip_extra}">{letter}</div>'
    return f'<div class="week-grid">{cells}</div>'


def _view_page_html(rows, kind, label, query_string):
    if kind == "simple":
        cols = [("Employee", None), ("Org Unit", "org"), ("Practice", "practice"),
                ("Sub Practice", "subPractice"), ("Designation", "designation"),
                ("Level", "level"), ("Client Location (State)", "clientLocationState"),
                ("Client Location (City)", "clientLocationCity"), ("Location", "location"),
                ("Reporting Manager", "reportingManager"), ("Date of Joining", "dateOfJoining")]
    elif kind == "exception":
        cols = [("Employee", None), ("Practice", "practice"), ("Sub Practice", "subPractice"), ("Designation", "designation"),
                ("Level", "level"), ("Location", "location"), ("Reporting Manager", "reportingManager"),
            ("Date of Joining", "dateOfJoining"), ("WFH Start", "wfhStart"), ("WFH End", "wfhEnd"), ("Employee Reason For Wfh", "reason"),
            ("Employee Detailed Reason", "detailedReason")]
    else:
        cols = [("Employee", None), ("Practice", "practice"), ("Sub Practice", "subPractice"), ("Designation", "designation"),
                ("Level", "level"), ("Project", "project"), ("Business Unit", "businessUnit"),
                ("Location", "location"), ("Reporting Manager", "reportingManager"), ("Date of Joining", "dateOfJoining"), ("Week", "days"), ("Total Hours", "totalWorkedHours"),
                ("Office Days", "officeDaysCount"), ("Leave Days", "leaveDaysCount"), ("Status", "subStatus")]

    def cell(e, field):
        if field is None:
            return f'<div class="emp-name">{e.get("name","")}</div><div class="emp-id">{e.get("id","")}</div>'
        if field == "subStatus":
            v = e.get("subStatus", "")
            badge_class = {"compliant": "compliant", "flagged": "flagged"}.get(v, "flagged")
            label_map = {"compliant": "Compliant", "flagged": "Non-compliant"}
            return f'<span class="badge {badge_class}">{label_map.get(v, v)}</span>'
        if field == "days":
            return _day_cell_html(e.get("days", []))
        if field == "totalWorkedHours":
            return _total_hours_from_days(e.get("days", []))
        return str(e.get(field, ""))

    rows_html = "".join(
        "<tr>" + "".join(f"<td>{cell(e, f)}</td>" for _, f in cols) + "</tr>"
        for e in rows
    ) or f'<tr><td colspan="{len(cols)}" style="text-align:center;color:var(--ink-soft);padding:30px;">No employees match this view.</td></tr>'

    col_filter_icon = (
        '<svg width="11" height="11" viewBox="0 0 24 24" fill="none" stroke="currentColor" '
        'stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round">'
        '<polygon points="22 3 2 3 10 12.46 10 19 14 21 14 12.46 22 3"></polygon></svg>'
    )

    header_html = "<tr>" + "".join(
        f'''<th data-col-idx="{idx}">
            <div class="th-row">
              <span class="sortable-th" title="Sort">{h} <span class="sort-icon">↕</span></span>
              <span class="col-filter-icon" title="Filter {h}">{col_filter_icon}</span>
            </div>
            <input type="text" class="col-filter-input" placeholder="Filter...">
          </th>'''
        for idx, (h, _) in enumerate(cols)
    ) + "</tr>"

    return (
        f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8"><title>{label} — Work Mode Policy Compliance</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@500;600;700&family=Inter:wght@400;500;600&family=IBM+Plex+Mono:wght@400;500&display=swap" rel="stylesheet">
<link rel="stylesheet" href="styles.css">
</head><body>
<div class="wrap" style="padding-top:30px;">
  <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:18px;flex-wrap:wrap;gap:12px;">
    <h1 class="display" style="font-size:22px;margin:0;">{label} <span class="count" style="font-size:14px;">({len(rows)})</span></h1>
    <a class="export-btn" href="export.xlsx?{query_string}" style="text-decoration:none;">Export to Excel</a>
  </div>
    <div class="table-vscroll"><div class="table-scroll"><table class="roster" id="detailRosterTable"><thead>{header_html}</thead><tbody>{rows_html}</tbody></table></div></div>
</div>
"""
                + """
<script>
    (function(){
        const table = document.getElementById('detailRosterTable');
        if(!table) return;

        let sortCol = null;
        let sortDir = 'asc';

        function sortRowsByColumn(colIdx){
            const tbody = table.querySelector('tbody');
            const rows = Array.from(tbody.querySelectorAll('tr'));
            if(!rows.length) return;

            if(sortCol === colIdx){
                sortDir = sortDir === 'asc' ? 'desc' : 'asc';
            } else {
                sortCol = colIdx;
                sortDir = 'asc';
            }

            rows.sort((a, b) => {
                const at = (a.children[colIdx]?.textContent || '').trim().toLowerCase();
                const bt = (b.children[colIdx]?.textContent || '').trim().toLowerCase();
                const an = Number(at.replace(/[^0-9.-]/g, ''));
                const bn = Number(bt.replace(/[^0-9.-]/g, ''));
                const bothNumeric = at !== '' && bt !== '' && !Number.isNaN(an) && !Number.isNaN(bn);
                if (bothNumeric) return an - bn;
                return at.localeCompare(bt);
            });

            if(sortDir === 'desc') rows.reverse();
            rows.forEach(r => tbody.appendChild(r));
            applyFilters();
            refreshSortIcons();
        }

        function refreshSortIcons(){
            table.querySelectorAll('thead th').forEach((th, idx) => {
                const icon = th.querySelector('.sort-icon');
                if(!icon) return;
                if(sortCol !== idx){
                    icon.textContent = '↕';
                    icon.classList.remove('active');
                    return;
                }
                icon.textContent = sortDir === 'asc' ? '▲' : '▼';
                icon.classList.add('active');
            });
        }

        function applyFilters(){
            const active = [];
            table.querySelectorAll('thead th[data-col-idx]').forEach(th => {
                const idx = parseInt(th.getAttribute('data-col-idx'), 10);
                const input = th.querySelector('.col-filter-input');
                const q = input ? input.value.trim().toLowerCase() : '';
                if(q) active.push({ idx, q });
            });

            table.querySelectorAll('tbody tr').forEach(tr => {
                let show = true;
                for(const f of active){
                    const text = (tr.children[f.idx]?.textContent || '').toLowerCase();
                    if(!text.includes(f.q)){ show = false; break; }
                }
                tr.style.display = show ? '' : 'none';
            });
        }

        table.querySelectorAll('thead th[data-col-idx]').forEach(th => {
            const idx = parseInt(th.getAttribute('data-col-idx'), 10);
            const sortable = th.querySelector('.sortable-th');
            const filterIcon = th.querySelector('.col-filter-icon');
            const filterInput = th.querySelector('.col-filter-input');

            if(sortable){
                sortable.addEventListener('click', () => sortRowsByColumn(idx));
            }

            if(filterIcon && filterInput){
                filterIcon.addEventListener('click', () => {
                    const showing = filterInput.classList.toggle('show');
                    if(showing){
                        filterInput.focus();
                    } else if(filterInput.value){
                        filterInput.value = '';
                        applyFilters();
                    }
                });
                filterInput.addEventListener('input', applyFilters);
            }
        });
    })();
</script>
</body></html>"""
    )


@app.route("/view")
def view_employees():
    rows, kind, label, error = _resolve_view_data()
    if error:
        return f'<div style="font-family:sans-serif;padding:40px;">{error}</div>', 200
    return _view_page_html(rows, kind, label, request.query_string.decode())


@app.route("/hr/view")
def view_employees_prefixed():
    return view_employees()


@app.route("/export.xlsx")
def export_xlsx():
    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    rows, kind, label, error = _resolve_view_data()
    if error:
        return error, 400

    if kind == "simple":
        columns = ["id", "name", "org", "practice", "subPractice", "designation", "level", "hrbp", "clientLocationState", "clientLocationCity", "location", "reportingManager", "dateOfJoining", "businessUnit"]
        headers = ["Employee ID", "Name", "Org Unit", "Practice", "Sub Practice", "Designation", "Level", "HRBP", "Client Location (State)", "Client Location (City)", "Location", "Reporting Manager", "Date of Joining", "Business Unit"]
        records = [{h: row.get(c, "") for h, c in zip(headers, columns)} for row in rows]
    elif kind == "exception":
        columns = ["id", "name", "practice", "subPractice", "designation", "level", "hrbp", "location", "reportingManager", "dateOfJoining", "businessUnit", "wfhStart", "wfhEnd", "reason", "detailedReason"]
        headers = ["Employee ID", "Name", "Practice", "Sub Practice", "Designation", "Level", "HRBP", "Location", "Reporting Manager", "Date of Joining", "Business Unit", "WFH Start", "WFH End", "Employee Reason For Wfh", "Employee Detailed Reason"]
        records = [{h: row.get(c, "") for h, c in zip(headers, columns)} for row in rows]
    else:
        columns = ["id", "name", "practice", "subPractice", "designation", "level", "hrbp",
                   "project", "businessUnit", "location", "reportingManager", "dateOfJoining", "officeDaysCount", "leaveDaysCount", "subStatus"]
        headers = ["Employee ID", "Name", "Practice", "Sub Practice", "Designation", "Level", "HRBP",
                   "Project", "Business Unit", "Location", "Reporting Manager", "Date of Joining", "Office Days", "Leave Days", "Status"]
        records = []
        for row in rows:
            rec = {h: row.get(c, "") for h, c in zip(headers, columns)}
            days = row.get("days", [])
            office_dates = ", ".join(d["date"] for d in days if d.get("state") == "office")
            leave_dates = ", ".join(d["date"] for d in days if d.get("state") == "leave")
            office_time_details = ", ".join(
                f'{d.get("date", "")}: in {d.get("firstIn", "")}'
                + (f', out {d.get("lastOut", "")}' if d.get("lastOut") else "")
                + (f' ({d.get("hours", "")})' if d.get("hours") else "")
                for d in days if d.get("state") == "office"
            )
            rec["Total Worked Hours"] = _total_hours_from_days(days)
            rec["Office Day Dates"] = office_dates
            rec["Leave Day Dates"] = leave_dates
            rec["Office Time Details"] = office_time_details
            records.append(rec)
        headers = headers + ["Total Worked Hours", "Office Day Dates", "Leave Day Dates", "Office Time Details"]

    df = pd.DataFrame(records, columns=headers)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Export", index=False)
        ws = writer.sheets["Export"]

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0B5D52", end_color="0B5D52", fill_type="solid")
        header_align = Alignment(horizontal="left", vertical="center", wrap_text=False)

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        ws.freeze_panes = "A2"

        for col_cells in ws.columns:
            max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
            col_letter = get_column_letter(col_cells[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 50)

        ws.row_dimensions[1].height = 20

    buffer.seek(0)
    safe_label = "".join(c if c.isalnum() else "_" for c in label)[:50]
    filename = f"{safe_label}_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(buffer, as_attachment=True, download_name=filename,
                      mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/hr/export.xlsx")
def export_xlsx_prefixed():
    return export_xlsx()


if __name__ == "__main__":
    app.run(host=config.HOST, port=config.PORT, debug=config.DEBUG)